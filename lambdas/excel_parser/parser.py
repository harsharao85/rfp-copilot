"""RFP workbook parser.

Problem: incoming RFP Excels come in wildly inconsistent formats:
- Questions in column A, column B, or spread across a question block
- Answer cells to the right of questions or directly below
- Merged cells for multi-line questions
- Section headers inline with questions
- Instructional paragraphs mixed in
- Multi-sheet workbooks with a master sheet + detail sheets

Strategy:
1. Walk every non-empty cell on every sheet.
2. Heuristic pre-pass detects the most likely (question-column, answer-column)
   pair per sheet by looking at text length distributions and presence of
   trailing '?' or numbering patterns (e.g., '1.', 'Q1', 'SEC-001').
3. For cells we're uncertain about, a Haiku 4.5 classifier labels each as
   question | section_header | instruction | answer_target | metadata | unknown.
4. Merged cells are collapsed to their anchor coordinate and treated as one logical cell.
5. The writer later uses the `answer_cell` CellRef to update cells in place
   without rebuilding the workbook (preserving formatting, formulas, validations).

The classifier step is deferred to a separate Lambda in prod to keep
the parser deterministic (no LLM) for ~80% of well-formed RFPs. Only
if heuristics flag low confidence do we call Haiku.
"""
from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from shared.models import CellRef, Question

# Patterns used for heuristic detection
QUESTION_NUMBER_RE = re.compile(
    r"^\s*("
    r"Q\d+|"                  # Q1, Q12
    r"\d+[\.\)]|"             # 1., 1)
    r"\d+\.\d+[\.\)]?|"       # 1.1, 1.2.3
    r"[A-Z]{2,6}-\d+|"        # SEC-001, ACC-12
    r"[A-Z]\.\d+|"            # A.1
    r"Question\s+\d+"
    r")\s",
    re.IGNORECASE,
)
SECTION_HEADER_RE = re.compile(
    r"^\s*("
    r"Section\s+\d+|"
    r"Part\s+[A-Z]|"
    r"[A-Z\s]{4,}$|"            # ALL CAPS short line
    r"\d+\.\s+[A-Z][a-z]+|"     # "1. Security"
    r"[A-Z]\.\s+[A-Z][a-z]+|"   # "A. Corporate"
    r"[A-Z]\d+\.\s+[A-Z][a-z]+" # "A1. Corporate"
    r")",
)
# Keep this tight. Anything matched here is CLASSIFIED AS NON-QUESTION,
# so a false positive silently drops a real question. Each keyword is
# anchored to the start of the cell (see _looks_like_instruction) to
# avoid swallowing questions that merely *contain* these phrases
# (e.g., "How do you respond to insider threats?").
INSTRUCTION_KEYWORDS = (
    "please respond", "please provide", "please answer", "please complete",
    "instructions:", "note:", "guidance:", "please fill in",
)


@dataclass
class Candidate:
    """A cell we think might be a question."""
    sheet: str
    row: int
    col: int
    text: str
    has_number: bool
    ends_with_q: bool
    length: int


def _cell_is_blank(c: Cell) -> bool:
    return c.value is None or (isinstance(c.value, str) and c.value.strip() == "")


def _looks_like_section_header(text: str) -> bool:
    if len(text) > 80:
        return False
    # Must be multi-word — filters out single-token IDs like "SEC-001"
    # and "Q1" which are question identifiers, not section headers.
    words = text.split()
    if len(words) < 2:
        return False
    if SECTION_HEADER_RE.match(text):
        return True
    # All-caps multi-word short lines with no trailing question mark
    if text == text.upper() and len(words) <= 6 and "?" not in text:
        return True
    return False


def _looks_like_instruction(text: str) -> bool:
    """Start-anchored match: instruction cells begin with the keyword.
    Questions that happen to contain the keyword mid-sentence are not
    misclassified."""
    t = text.strip().lower()
    return any(t.startswith(kw) for kw in INSTRUCTION_KEYWORDS)


def _classify_question_candidate(c: Candidate) -> str:
    """Heuristic label: question | section_header | instruction | metadata."""
    text = c.text.strip()
    if len(text) < 8:
        return "metadata"
    if _looks_like_instruction(text):
        return "instruction"
    if _looks_like_section_header(text):
        return "section_header"
    if c.ends_with_q or c.has_number or len(text) > 30:
        return "question"
    return "metadata"


def _merged_cell_anchor(ws: Worksheet, row: int, col: int) -> tuple[int, int]:
    """If (row, col) is inside a merged range, return the anchor coordinate.
    Otherwise return (row, col) unchanged."""
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row
                and merged.min_col <= col <= merged.max_col):
            return merged.min_row, merged.min_col
    return row, col


def _detect_answer_column(ws: Worksheet, question_col: int) -> int:
    """Heuristic: answer column is typically the next column to the right
    that is mostly empty in rows containing questions. If no such column
    exists, we insert one at question_col + 1 when writing output."""
    max_col = ws.max_column
    for candidate in range(question_col + 1, max_col + 2):
        empties = 0
        checked = 0
        for row in ws.iter_rows(min_col=candidate, max_col=candidate, values_only=True):
            checked += 1
            if row[0] is None or (isinstance(row[0], str) and row[0].strip() == ""):
                empties += 1
            if checked > 30:
                break
        if checked == 0 or empties / checked > 0.5:
            return candidate
    return question_col + 1


def _detect_question_column(ws: Worksheet) -> int:
    """Pick the column with the highest density of question-shaped text."""
    scores: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            if len(text) < 10:
                continue
            c = Candidate(
                sheet=ws.title,
                row=cell.row,
                col=cell.column,
                text=text,
                has_number=bool(QUESTION_NUMBER_RE.match(text)),
                ends_with_q="?" in text,
                length=len(text),
            )
            if _classify_question_candidate(c) == "question":
                scores[cell.column] = scores.get(cell.column, 0) + 1
    if not scores:
        return 1  # default to column A
    return max(scores.items(), key=lambda kv: kv[1])[0]


def parse_workbook(path: str | Path, job_id: str | None = None) -> tuple[str, list[Question]]:
    """Parse an RFP workbook into a normalized list of Questions.

    Returns (job_id, questions). Caller is responsible for persisting to DynamoDB.
    """
    job_id = job_id or str(uuid.uuid4())
    wb = load_workbook(filename=str(path), data_only=False, keep_vba=False)

    questions: list[Question] = []
    current_section: str | None = None

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue

        q_col = _detect_question_column(ws)
        a_col = _detect_answer_column(ws, q_col)

        for row_cells in ws.iter_rows():
            # First pass across the row: does any cell look like a section header?
            # Section headers often live in column A (merged across the row) while
            # questions live in column B. We want to update current_section when
            # we see the header, regardless of which column it's in.
            for cell in row_cells:
                if _cell_is_blank(cell) or not isinstance(cell.value, str):
                    continue
                anchor_row, anchor_col = _merged_cell_anchor(ws, cell.row, cell.column)
                if (anchor_row, anchor_col) != (cell.row, cell.column):
                    continue
                text = cell.value.strip()
                if _looks_like_section_header(text):
                    current_section = text
                    break  # don't double-count within this row

            # Second pass: look for a question in the detected question column
            for cell in row_cells:
                if _cell_is_blank(cell) or not isinstance(cell.value, str):
                    continue
                anchor_row, anchor_col = _merged_cell_anchor(ws, cell.row, cell.column)
                if (anchor_row, anchor_col) != (cell.row, cell.column):
                    continue
                if cell.column != q_col:
                    continue

                text = cell.value.strip()
                candidate = Candidate(
                    sheet=ws.title,
                    row=cell.row,
                    col=cell.column,
                    text=text,
                    has_number=bool(QUESTION_NUMBER_RE.match(text)),
                    ends_with_q="?" in text,
                    length=len(text),
                )
                label = _classify_question_candidate(candidate)

                if label != "question":
                    continue

                questions.append(
                    Question(
                        job_id=job_id,
                        question_id=f"q-{len(questions) + 1:04d}",
                        text=text,
                        section=current_section,
                        context=None,
                        answer_cell=CellRef(
                            sheet=ws.title,
                            coordinate=f"{get_column_letter(a_col)}{cell.row}",
                        ),
                        confidence_cell=CellRef(
                            sheet=ws.title,
                            coordinate=f"{get_column_letter(a_col + 1)}{cell.row}",
                        ),
                    )
                )

    return job_id, questions
