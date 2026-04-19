"""Output Excel writer.

Critical invariant: we open the ORIGINAL workbook and mutate only
the specific answer + confidence cells. We do not rebuild the sheet.
openpyxl preserves formulas, styles, data validation, hidden columns,
and conditional formatting. Additions:
  - Answer text in answer_cell
  - Confidence score in confidence_cell
  - Cell fill color tied to tier (green/amber/red)
  - Cell comment on the answer cell listing citations + rule triggers
  - A new Summary sheet appended at the end

Colors match the v0.4 plan §5.8 convention:
  green  = C6EFCE
  amber  = FFEB9C
  red    = FFC7CE
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from shared.models import FinalAnswer, Tier

TIER_FILLS = {
    Tier.GREEN: PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    Tier.AMBER: PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
    Tier.RED:   PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
}

TIER_COMMENT_PREFIX = {
    Tier.GREEN: "GREEN — auto-draft ready for rep review.",
    Tier.AMBER: "AMBER — SME review required.",
    Tier.RED:   "RED — SME/compliance/commercial review REQUIRED before submission.",
}


def _comment_body(ans: FinalAnswer) -> str:
    lines = [TIER_COMMENT_PREFIX[ans.tier]]
    lines.append("")
    lines.append(f"Confidence: {ans.raw_confidence:.2f}")
    b = ans.confidence_breakdown
    lines.append(f"  H(prior)={b.h:.2f}  R(retrieval)={b.r:.2f}  C(coverage)={b.c:.2f}  F(fresh)={b.f:.2f}  G(guardrail)={b.g:.2f}")
    if ans.hard_rule_triggers:
        lines.append("")
        lines.append("Hard-rule triggers:")
        for t in ans.hard_rule_triggers:
            lines.append(f"  • {t}")
    if ans.citations:
        lines.append("")
        lines.append("Citations:")
        for cite in ans.citations[:5]:
            lines.append(f"  • {cite}")
    return "\n".join(lines)


def write_output(
    *,
    source_path: str | Path,
    dest_path: str | Path,
    answers: list[FinalAnswer],
    job_id: str,
) -> Path:
    """Write answers into a copy of the source workbook. Returns dest_path."""
    wb = load_workbook(filename=str(source_path))

    # Index answers by sheet so we can visit each sheet once
    by_sheet: dict[str, list[FinalAnswer]] = {}
    for ans in answers:
        by_sheet.setdefault(ans.answer_cell.sheet, []).append(ans)

    for sheet_name, sheet_answers in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for ans in sheet_answers:
            try:
                answer_cell = ws[ans.answer_cell.coordinate]
                confidence_cell = ws[ans.confidence_cell.coordinate]
            except (KeyError, ValueError):
                continue

            tier = ans.tier
            answer_cell.value = ans.answer_text
            answer_cell.alignment = Alignment(wrap_text=True, vertical="top")
            answer_cell.fill = TIER_FILLS[tier]
            answer_cell.comment = Comment(_comment_body(ans), "RFP Copilot")

            confidence_cell.value = round(ans.raw_confidence, 2)
            confidence_cell.fill = TIER_FILLS[tier]
            confidence_cell.alignment = Alignment(horizontal="center")

    # Append Summary sheet
    _write_summary_sheet(wb, answers=answers, job_id=job_id)

    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest_path))
    return dest_path


def _write_summary_sheet(wb: Any, *, answers: list[FinalAnswer], job_id: str) -> None:  # type: ignore[no-untyped-def]
    if "RFP Copilot Summary" in wb.sheetnames:
        del wb["RFP Copilot Summary"]
    summary = wb.create_sheet("RFP Copilot Summary")

    summary["A1"] = "RFP Copilot — Answer Summary"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = f"Job ID: {job_id}"
    summary["A3"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"
    summary["A5"] = "Tier distribution"
    summary["A5"].font = Font(bold=True)

    counts = {Tier.GREEN: 0, Tier.AMBER: 0, Tier.RED: 0}
    for a in answers:
        counts[a.tier] += 1

    row = 6
    for tier, n in counts.items():
        summary.cell(row=row, column=1, value=tier.value.upper()).fill = TIER_FILLS[tier]
        summary.cell(row=row, column=2, value=n)
        summary.cell(row=row, column=3, value=f"{(n / max(len(answers), 1) * 100):.1f}%")
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Total questions").font = Font(bold=True)
    summary.cell(row=row, column=2, value=len(answers))
    row += 2

    summary.cell(row=row, column=1, value="Review checklist").font = Font(bold=True)
    row += 1
    trigger_counts: dict[str, int] = {}
    for a in answers:
        for t in a.hard_rule_triggers:
            # Collapse unapproved_reference:<name> variants
            key = t.split(":", 1)[0] if ":" in t else t
            trigger_counts[key] = trigger_counts.get(key, 0) + 1
    for k, v in sorted(trigger_counts.items(), key=lambda kv: -kv[1]):
        summary.cell(row=row, column=1, value=k)
        summary.cell(row=row, column=2, value=v)
        row += 1

    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 14
