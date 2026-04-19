"""Create 3 demo RFP Excel files for the RFP Copilot demo.

Set A: demo_rfp_techops.xlsx    — Mostly Green  (~7G, 3A, 2R)
Set B: demo_rfp_compliance.xlsx — Mostly Amber  (~2G, 8A, 2R)
Set C: demo_rfp_enterprise.xlsx — Mostly Red    (~1G, 4A, 7R)

Each targets specific scoring outcomes based on the confidence composite
formula (0.45·H + 0.25·R + 0.15·C + 0.10·F + 0.05·G) and hard rules.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).parent.parent
OUT = ROOT / "data" / "incoming"


def _make_workbook(title: str, prospect: str, sections: list[tuple[str, list[tuple[str, str]]]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFP"

    header_font = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append([title])
    ws.append([f"Prospect: {prospect} (fictional)"])
    ws.append(["Sent: 2026-04-16"])
    ws.append(["Please respond to each question in the Answer column."])
    ws.append([])
    ws.append(["ID", "Question", "Answer", "Confidence"])
    for cell in ws[6]:
        cell.font = header_font

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 12

    for section_name, questions in sections:
        row = [section_name]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = section_fill
            cell.font = header_font
        for qid, qtext in questions:
            ws.append([qid, qtext, None, None])

    return wb


# ---------------------------------------------------------------------------
# Set A — TechOps RFP ("Mostly Green")
# Topic alignment: encryption/TLS, SSO/MFA/SCIM, DR/backup, incident response
# → these all have LibraryFeedback H-signal + S3 corpus passages → GREEN
# Amber: pen testing (compliance, corroboration_required), SLA uptime (forward-looking)
# Red: pricing (hard rule commercial), reference customer with expired approval
# ---------------------------------------------------------------------------
SET_A_SECTIONS = [
    ("A. Data Protection", [
        ("TECH-001", "How is customer data encrypted at rest? Specify algorithm and key management."),
        ("TECH-002", "How is customer data encrypted in transit? Which TLS versions are supported?"),
        ("TECH-003", "Do you support customer-managed encryption keys (BYOK)? Which KMS providers?"),
    ]),
    ("B. Identity & Access", [
        ("TECH-004", "Which SSO protocols do you support? (SAML 2.0, OIDC, OAuth 2.0)"),
        ("TECH-005", "Is MFA supported for all users? Which authentication factors are available?"),
        ("TECH-006", "Does your platform support SCIM 2.0 for automated user provisioning?"),
    ]),
    ("C. Availability & Recovery", [
        ("TECH-007", "Describe your disaster recovery architecture, RTO, and RPO targets."),
        ("TECH-008", "What are your backup procedures? How frequently are backups taken and tested?"),
        ("TECH-009", "What uptime SLA do you commit to? Will you provide SLA credits for downtime?"),
    ]),
    ("D. Security Operations", [
        ("TECH-010", "Describe your incident response process and customer notification timelines."),
        ("TECH-011", "How frequently do you conduct penetration tests? Who performs them?"),
    ]),
    ("E. Commercial", [
        ("TECH-012", "What is your per-seat pricing for 500 users on the Enterprise tier?"),
    ]),
]

# ---------------------------------------------------------------------------
# Set B — Compliance RFP ("Mostly Amber")
# Compliance questions require corroboration_required → max AMBER without SME sign-off
# Two hard-rule questions for RED demo
# One or two well-covered product questions for GREEN
# ---------------------------------------------------------------------------
SET_B_SECTIONS = [
    ("A. Certifications", [
        ("COMP-001", "What compliance certifications does your organization hold? Provide scope, issuing body, and last audit date."),
        ("COMP-002", "Describe your SOC 2 Type II audit scope and trust principles covered."),
        ("COMP-003", "Are you ISO 27001 certified? If yes, provide the certificate number and expiry date."),
        ("COMP-004", "What is your FedRAMP authorization status and anticipated authorization timeline?"),
    ]),
    ("B. Privacy & GDPR", [
        ("COMP-005", "Describe your GDPR compliance approach for EU data subjects' rights requests."),
        ("COMP-006", "Do you offer a Data Processing Agreement? Describe key terms."),
        ("COMP-007", "Provide a complete list of sub-processors and their processing locations."),
        ("COMP-008", "Can customers configure data residency to restrict data storage to specific regions?"),
    ]),
    ("C. Product Security", [
        ("COMP-009", "How do you detect and respond to insider threats?"),
        ("COMP-010", "What is your vulnerability disclosure and patching SLA?"),
    ]),
    ("D. Commercial", [
        ("COMP-011", "What are your volume discount tiers for multi-year contracts?"),
        ("COMP-012", "How does your product compare to Competitor X in terms of security posture?"),
    ]),
]

# ---------------------------------------------------------------------------
# Set C — Enterprise RFP ("Mostly Red/Amber")
# Designed to show what a rep sees when questions are commercially sensitive,
# forward-looking, or reference unapproved customers.
# ---------------------------------------------------------------------------
SET_C_SECTIONS = [
    ("A. Commercial Terms", [
        ("ENT-001", "Provide your complete Enterprise tier pricing schedule including add-on modules."),
        ("ENT-002", "What volume discounts apply at 1,000, 5,000, and 10,000 seats?"),
        ("ENT-003", "Describe the financial penalties your SLA includes for downtime breaches."),
        ("ENT-004", "Are custom contract terms negotiable? What is your standard MSA revision process?"),
        ("ENT-005", "What are your professional services rates for implementation and training?"),
    ]),
    ("B. Forward-Looking Commitments", [
        ("ENT-006", "When will you achieve FedRAMP High authorization? Can you commit to a delivery date?"),
        ("ENT-007", "Will you guarantee 99.99% uptime in a signed SLA for our contract period?"),
        ("ENT-008", "Can you guarantee data sovereignty — that our data will never leave EU jurisdiction?"),
        ("ENT-009", "What new security features will you deliver in the next 12 months?"),
    ]),
    ("C. Reference Customers", [
        ("ENT-010", "Provide a named reference from a Fortune 500 logistics company using your Enterprise tier."),
    ]),
    ("D. Competitive", [
        ("ENT-011", "Why is your security posture superior to Competitor Y? Provide a feature comparison."),
        ("ENT-012", "What incident SLA penalties do you offer that Competitor Y does not?"),
    ]),
]


def main() -> None:
    wb_a = _make_workbook(
        "TechOps Vendor Security Questionnaire",
        "TechOps Holdings",
        SET_A_SECTIONS,
    )
    wb_a.save(OUT / "demo_rfp_techops.xlsx")
    print(f"Created {OUT / 'demo_rfp_techops.xlsx'}")

    wb_b = _make_workbook(
        "Compliance & Privacy Assessment",
        "BlueSky Financial Services",
        SET_B_SECTIONS,
    )
    wb_b.save(OUT / "demo_rfp_compliance.xlsx")
    print(f"Created {OUT / 'demo_rfp_compliance.xlsx'}")

    wb_c = _make_workbook(
        "Enterprise Vendor Evaluation",
        "GlobalCorp Enterprises",
        SET_C_SECTIONS,
    )
    wb_c.save(OUT / "demo_rfp_enterprise.xlsx")
    print(f"Created {OUT / 'demo_rfp_enterprise.xlsx'}")


if __name__ == "__main__":
    main()
