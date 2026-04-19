"""Synthetic data generator for the RFP Redlining Copilot demo.

Produces all artifacts under rfp-copilot/data/:
- incoming/sample_rfp_acmesec.xlsx  (a realistic-shaped security questionnaire)
- historical/won_rfp_*.xlsx         (prior SME-approved answers — the reference corpus)
- seismic/content_cards.json        (Seismic content export mock)
- gong/transcripts.json             (Gong call transcript mock)
- whitepapers/*.md                  (policies, overviews, bridge letters)
- graph/customers.json              (customer reference approval status)
- graph/products.json               (product / topic / competency graph seed)
- corpus/compliance/*.json          (primary source sidecars — compliance_store)
- corpus/product-docs/*.json        (primary source sidecars — product_docs)
- corpus/prior-rfps/*.json          (prior RFP answer sidecars — includes 3 stale, 1 fresh)

Fully fictitious. No real customer names.

Run:   python scripts/generate_synthetic_data.py
Reset: remove data/ artifacts; script is idempotent.
"""
from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RNG = Random(42)  # deterministic demo
ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

# --------------------------------------------------------------------------
# Incoming RFP: a security questionnaire with ~30 questions of varied shape
# --------------------------------------------------------------------------

INCOMING_RFP_QUESTIONS: list[dict[str, str]] = [
    # Section A: Corporate & compliance
    {"section": "A. Corporate & Compliance", "id": "SEC-001", "q": "What compliance certifications does your organization currently hold? Please list certification name, scope, issuing body, and most recent audit date."},
    {"section": "A. Corporate & Compliance", "id": "SEC-002", "q": "Describe your SOC 2 Type II audit scope and the trust principles covered."},
    {"section": "A. Corporate & Compliance", "id": "SEC-003", "q": "Are you ISO 27001 certified? If yes, provide certificate number and expiration."},
    {"section": "A. Corporate & Compliance", "id": "SEC-004", "q": "Provide a reference customer of similar size in our industry (financial services, 5,000+ employees) who has deployed your platform in production."},
    {"section": "A. Corporate & Compliance", "id": "SEC-005", "q": "Describe your approach to GDPR compliance for EU data subjects."},

    # Section B: Data protection
    {"section": "B. Data Protection", "id": "SEC-006", "q": "How is customer data encrypted at rest? Specify algorithm, key length, and key management approach."},
    {"section": "B. Data Protection", "id": "SEC-007", "q": "How is customer data encrypted in transit? Specify protocol versions and cipher suites supported."},
    {"section": "B. Data Protection", "id": "SEC-008", "q": "Do you support customer-managed encryption keys (BYOK / HYOK)? If yes, which key-management services are supported?"},
    {"section": "B. Data Protection", "id": "SEC-009", "q": "Describe your approach to data classification and handling of sensitive data."},
    {"section": "B. Data Protection", "id": "SEC-010", "q": "What is your data retention and deletion policy? Include timelines for customer-initiated deletion."},

    # Section C: Identity & access
    {"section": "C. Identity & Access", "id": "SEC-011", "q": "Which SSO protocols do you support? (SAML 2.0, OIDC, OAuth 2.0, etc.)"},
    {"section": "C. Identity & Access", "id": "SEC-012", "q": "Do you support multi-factor authentication for all users? Which factors are supported?"},
    {"section": "C. Identity & Access", "id": "SEC-013", "q": "Describe your approach to role-based access control and least-privilege enforcement."},
    {"section": "C. Identity & Access", "id": "SEC-014", "q": "Do you support SCIM for automated user provisioning and deprovisioning?"},

    # Section D: Availability & resilience
    {"section": "D. Availability & Resilience", "id": "SEC-015", "q": "What is your SLA commitment for platform uptime? Provide monthly and annual figures."},
    {"section": "D. Availability & Resilience", "id": "SEC-016", "q": "Describe your disaster recovery capabilities including RTO and RPO targets."},
    {"section": "D. Availability & Resilience", "id": "SEC-017", "q": "How are backups performed, stored, and tested?"},

    # Section E: Operations & incident response
    {"section": "E. Operations & Incident Response", "id": "SEC-018", "q": "Describe your incident response process, including notification timelines to affected customers."},
    {"section": "E. Operations & Incident Response", "id": "SEC-019", "q": "Do you conduct penetration testing? How often, by whom, and are reports available to customers?"},
    {"section": "E. Operations & Incident Response", "id": "SEC-020", "q": "Describe your vulnerability management program, including SLAs for critical patches."},
    {"section": "E. Operations & Incident Response", "id": "SEC-021", "q": "How do you detect and respond to insider threats?"},

    # Section F: Application security
    {"section": "F. Application Security", "id": "SEC-022", "q": "Describe your secure software development lifecycle (SSDLC) practices."},
    {"section": "F. Application Security", "id": "SEC-023", "q": "Do you perform static and dynamic code analysis? Which tools are in use?"},
    {"section": "F. Application Security", "id": "SEC-024", "q": "How do you manage dependencies and third-party library vulnerabilities?"},

    # Section G: Privacy & legal
    {"section": "G. Privacy & Legal", "id": "SEC-025", "q": "Provide your standard Data Processing Agreement (DPA) and list sub-processors."},
    {"section": "G. Privacy & Legal", "id": "SEC-026", "q": "In which regions is customer data physically stored? Is data residency configurable by customer?"},
    {"section": "G. Privacy & Legal", "id": "SEC-027", "q": "Describe your process for responding to customer data subject access requests (DSARs)."},

    # Section H: Commercial
    {"section": "H. Commercial", "id": "SEC-028", "q": "Provide standard per-user pricing for your Enterprise tier, including any volume discounts available at 5,000+ seats."},
    {"section": "H. Commercial", "id": "SEC-029", "q": "What is your roadmap for FedRAMP Moderate authorization, and when will it be available?"},
    {"section": "H. Commercial", "id": "SEC-030", "q": "Are you willing to commit to a 99.99% uptime SLA with financial penalties for breach?"},
]


def _make_rfp_workbook(
    dest: Path,
    *,
    title: str,
    prospect: str,
    questions: list[dict[str, str]],
    with_answers: bool = False,
    win_loss: str | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Questionnaire"

    # Header block (mixed formatting — the parser must ignore these)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Prospect: {prospect}"
    ws["A3"] = f"Sent: {datetime.now(timezone.utc).date().isoformat()}"
    ws["A4"] = "Please respond to each question in the Answer column. Additional context may be provided in the Notes column."
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A4:D4")
    ws.row_dimensions[4].height = 30

    # Column headers
    ws["A6"] = "ID"
    ws["B6"] = "Question"
    ws["C6"] = "Answer"
    ws["D6"] = "Confidence"
    for col in "ABCD":
        ws[f"{col}6"].font = Font(bold=True)
        ws[f"{col}6"].fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 12

    row = 7
    current_section: str | None = None
    for q in questions:
        if q["section"] != current_section:
            current_section = q["section"]
            ws.cell(row=row, column=1, value=current_section).font = Font(bold=True, italic=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
            row += 1

        ws.cell(row=row, column=1, value=q["id"])
        cell_b = ws.cell(row=row, column=2, value=q["q"])
        cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        if with_answers and "a" in q:
            cell_c = ws.cell(row=row, column=3, value=q["a"])
            cell_c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 48
        row += 1

    if with_answers and win_loss:
        # Meta row for historical RFPs to help Kendra metadata tagging
        meta_row = row + 2
        ws.cell(row=meta_row, column=1, value="WIN/LOSS").font = Font(bold=True)
        ws.cell(row=meta_row, column=2, value=win_loss)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))


# --------------------------------------------------------------------------
# Historical RFPs (reference corpus) — condensed, with SME-approved answers
# --------------------------------------------------------------------------

HISTORICAL_Q_AND_A: list[dict[str, str]] = [
    {
        "section": "Data Protection",
        "id": "HIST-001",
        "q": "How is customer data encrypted at rest?",
        "a": "All customer data is encrypted at rest using AES-256-GCM with keys managed by AWS KMS. Customer-managed keys (BYOK) are supported for Enterprise customers via AWS KMS CMKs in the customer's account, referenced by key-policy. Key rotation is enabled by default on an annual cadence.",
    },
    {
        "section": "Data Protection",
        "id": "HIST-002",
        "q": "How is data encrypted in transit?",
        "a": "All data in transit is encrypted with TLS 1.2 or TLS 1.3. We require Perfect Forward Secrecy (ECDHE cipher suites), disable SSLv3 and TLS 1.0/1.1 at the load balancer, and publish HSTS headers with a one-year max-age.",
    },
    {
        "section": "Identity & Access",
        "id": "HIST-003",
        "q": "Which SSO protocols are supported?",
        "a": "We support SAML 2.0 and OIDC as identity providers for customer-initiated SSO. OAuth 2.0 Authorization Code flow with PKCE is used for our native integrations. SCIM 2.0 is supported for automated user provisioning.",
    },
    {
        "section": "Identity & Access",
        "id": "HIST-004",
        "q": "Is MFA supported? Which factors?",
        "a": "MFA is supported for all user tiers and can be enforced by administrator policy. Supported factors: TOTP authenticator apps, WebAuthn/FIDO2 security keys, and SMS (not recommended for high-assurance contexts). Hardware keys are required for administrative roles in our Enterprise tier.",
    },
    {
        "section": "Availability",
        "id": "HIST-005",
        "q": "Describe DR capabilities and RTO/RPO targets.",
        "a": "Our platform is deployed active-active across two AWS regions with synchronous replication of critical metadata and asynchronous replication (< 60 seconds) for bulk customer data. Target RTO is under four hours and target RPO is under fifteen minutes. DR exercises are conducted twice annually.",
    },
    {
        "section": "Operations",
        "id": "HIST-006",
        "q": "Describe your incident response process.",
        "a": "Security incidents are triaged by our 24x7 Security Operations Center within 15 minutes of detection. A named Incident Commander coordinates response following the NIST 800-61 framework. Customers affected by a confirmed security incident are notified within 72 hours, with an initial advisory containing scope, impact, and remediation status.",
    },
    {
        "section": "Privacy",
        "id": "HIST-007",
        "q": "Provide your DPA and list of sub-processors.",
        "a": "Our standard Data Processing Agreement is available on our Trust Center and incorporates the EU Standard Contractual Clauses. Our sub-processor list is maintained publicly and updated at least 30 days before any change. Sub-processors are audited annually against our third-party risk program.",
    },
    {
        "section": "Privacy",
        "id": "HIST-008",
        "q": "Where is customer data stored? Is residency configurable?",
        "a": "Customer data is stored in AWS regions elected at tenant creation. Supported regions: us-east-1, us-west-2, eu-west-1, eu-central-1, ap-southeast-2. Data does not leave the elected region for operational purposes. Cross-region replication is available as an opt-in feature for disaster recovery.",
    },
    {
        "section": "Application Security",
        "id": "HIST-009",
        "q": "Describe your SSDLC practices.",
        "a": "We follow a security development lifecycle aligned with OWASP SAMM. All code changes require peer review and pass automated SAST (Semgrep), SCA (Snyk), and container image scanning before merge. Threat modeling is performed for every new service using STRIDE. Annual penetration testing is conducted by an independent third party; executive summaries are available under NDA.",
    },
    {
        "section": "Application Security",
        "id": "HIST-010",
        "q": "How do you manage third-party library vulnerabilities?",
        "a": "We maintain a software bill of materials (SBOM) for every deployable artifact. Vulnerabilities are triaged daily with SLAs of 24 hours for critical, 7 days for high, and 30 days for medium. Patch status is visible in our internal risk register and summarized quarterly to customers under NDA.",
    },
]

WON_RFPS: list[tuple[str, str, list[dict[str, str]], str]] = [
    (
        "won_rfp_acme_financial_2025",
        "Acme Financial (fictional prospect, closed won 2025-Q3)",
        HISTORICAL_Q_AND_A[:6],
        "WON",
    ),
    (
        "won_rfp_bluebird_insurance_2026",
        "Bluebird Insurance (fictional prospect, closed won 2026-Q1)",
        HISTORICAL_Q_AND_A[3:],
        "WON",
    ),
    (
        "lost_rfp_globex_2025",
        "Globex (fictional prospect, closed lost 2025-Q4)",
        HISTORICAL_Q_AND_A[:5],
        "LOST",
    ),
]

# --------------------------------------------------------------------------
# Seismic + Gong mocks
# --------------------------------------------------------------------------

SEISMIC_CARDS: list[dict[str, str]] = [
    {
        "card_id": "seismic-card-001",
        "title": "Enterprise Security Overview One-Pager",
        "body": "High-level customer-facing overview: SOC 2 Type II (Trust Principles: Security, Availability, Confidentiality), ISO 27001, active FedRAMP Moderate authorization in progress. Encryption at rest AES-256-GCM, in transit TLS 1.2+. MFA mandatory for admin roles.",
        "updated_at": (datetime.now(timezone.utc) - timedelta(days=40)).date().isoformat(),
        "owner": "Product Marketing",
    },
    {
        "card_id": "seismic-card-002",
        "title": "FAQ — Reference Customer Stories (Financial Services)",
        "body": "Approved public references in the financial services vertical include logos listed in our public customer page. Manager must consult the Reference Customer List in Neptune before naming any customer in an outbound document.",
        "updated_at": (datetime.now(timezone.utc) - timedelta(days=20)).date().isoformat(),
        "owner": "Sales Enablement",
    },
    {
        "card_id": "seismic-card-003",
        "title": "Data Residency Configuration Matrix",
        "body": "Tenant creation supports region selection among us-east-1, us-west-2, eu-west-1, eu-central-1, ap-southeast-2. Cross-region replication is an opt-in feature for DR. Residency is enforced at the storage layer via KMS CMK scoping.",
        "updated_at": (datetime.now(timezone.utc) - timedelta(days=110)).date().isoformat(),
        "owner": "Security Engineering",
    },
    {
        "card_id": "seismic-card-004",
        "title": "SSDLC Summary Deck",
        "body": "Secure SDLC aligned to OWASP SAMM. SAST: Semgrep. SCA: Snyk. Container scanning: Trivy + ECR scan. Threat modeling: STRIDE per service. Annual third-party pentest. SBOM generated per build.",
        "updated_at": (datetime.now(timezone.utc) - timedelta(days=180)).date().isoformat(),
        "owner": "Application Security",
    },
    {
        "card_id": "seismic-card-005",
        "title": "Pricing — Internal Only (DO NOT SHARE)",
        "body": "Enterprise tier list price is $X per user per month with standard volume tiers. All pricing must be validated by Deal Desk before appearing in customer-facing documents.",
        "updated_at": (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat(),
        "owner": "Deal Desk",
    },
]

GONG_TRANSCRIPTS: list[dict[str, object]] = [
    {
        "call_id": "gong-call-001",
        "title": "Discovery call — prospect asks about FedRAMP timeline",
        "date": (datetime.now(timezone.utc) - timedelta(days=14)).date().isoformat(),
        "participants": ["AE: Jordan Avery", "SE: Priya Menon", "Prospect: VP InfoSec"],
        "snippets": [
            "Prospect asked when FedRAMP Moderate authorization will be live. Priya responded that the ATO assessment is in progress and she would not commit to a specific date in writing.",
            "Prospect pushed for an SLA with financial penalties; Jordan redirected to standard 99.9% SLA and noted financial penalties are negotiated at contract.",
        ],
    },
    {
        "call_id": "gong-call-002",
        "title": "Technical deep-dive — encryption and key management",
        "date": (datetime.now(timezone.utc) - timedelta(days=22)).date().isoformat(),
        "participants": ["SE: Priya Menon", "Prospect: Security Architect"],
        "snippets": [
            "Discussed BYOK via AWS KMS CMK in the customer's account. Confirmed that key deletion by the customer results in data becoming unreadable within the SLA window.",
            "Confirmed that all data-in-transit uses TLS 1.2 minimum, with HSTS, and that TLS 1.0/1.1 have been deprecated end-to-end.",
        ],
    },
    {
        "call_id": "gong-call-003",
        "title": "Procurement call — DPA and sub-processors",
        "date": (datetime.now(timezone.utc) - timedelta(days=9)).date().isoformat(),
        "participants": ["AE: Jordan Avery", "Legal: Sam Okafor", "Prospect: Procurement Lead"],
        "snippets": [
            "Walked through standard DPA. Confirmed EU SCCs are incorporated. Public sub-processor list is updated with 30-day advance notice of changes.",
            "Prospect asked to add a specific customer logo to the reference list. Jordan deferred pending approval from the reference customer.",
        ],
    },
]

# --------------------------------------------------------------------------
# Whitepapers (policies / overviews)
# --------------------------------------------------------------------------

WHITEPAPERS: dict[str, str] = {
    "security_overview": """# Security Overview

Our platform is designed around a defense-in-depth architecture with
mandatory controls at the network, compute, data, identity, and
application layers.

## Compliance
- SOC 2 Type II — annual audit, trust principles Security, Availability, Confidentiality
- ISO 27001 — certified
- FedRAMP Moderate — ATO in progress (no committed date)
- GDPR — EU SCCs incorporated into standard DPA

## Encryption
- At rest: AES-256-GCM, AWS KMS managed; BYOK available on Enterprise tier
- In transit: TLS 1.2+ with Perfect Forward Secrecy, HSTS, TLS 1.0/1.1 disabled
""",
    "soc2_bridge_letter": """# SOC 2 Type II Bridge Letter (Current)

This bridge letter covers the period between our most recent SOC 2
Type II report and the next scheduled audit. No material changes to
the control environment have occurred during the bridge period.

Scope: Security, Availability, Confidentiality trust principles.
Auditor: Independent third-party firm.
Next report: scheduled for issuance in Q3.
""",
    "data_protection_policy": """# Data Protection Policy

## Data classification
All customer data is classified at ingestion. Four tiers: Public,
Internal, Confidential, Restricted. Handling requirements escalate
by tier.

## Retention
Customer data is retained for the duration of the customer agreement
plus 90 days. Customer-initiated deletion is honored within 30 days;
cryptographic erasure of backups completes within the backup retention
window of 90 days.

## Residency
Tenant data is stored in the customer's elected AWS region and does
not leave that region for operational purposes.
""",
    "incident_response_overview": """# Incident Response Overview

Detection → Triage → Containment → Eradication → Recovery → Lessons
Learned. Aligned to NIST SP 800-61. 24x7 SOC. Customer notification
within 72 hours of a confirmed security incident affecting them.
""",
}

# --------------------------------------------------------------------------
# Graph seed data
# --------------------------------------------------------------------------

CUSTOMERS: list[dict[str, object]] = [
    {
        "customer_id": "cust-001",
        "name": "Northwind Capital",
        "industry": "financial_services",
        "employees": 8500,
        "contract_size_usd": 450000,
        "public_reference": True,
        "approval_expires": (datetime.now(timezone.utc) + timedelta(days=180)).date().isoformat(),
        "deployed_products": ["platform", "enterprise_tier"],
    },
    {
        "customer_id": "cust-002",
        "name": "Helix Biomedical",
        "industry": "healthcare",
        "employees": 3200,
        "contract_size_usd": 180000,
        "public_reference": True,
        "approval_expires": (datetime.now(timezone.utc) + timedelta(days=45)).date().isoformat(),
        "deployed_products": ["platform"],
    },
    {
        "customer_id": "cust-003",
        "name": "Aurora Federal Credit Union",
        "industry": "financial_services",
        "employees": 5400,
        "contract_size_usd": 320000,
        "public_reference": False,  # NDA — cannot be named
        "approval_expires": None,
        "deployed_products": ["platform", "enterprise_tier"],
    },
    {
        "customer_id": "cust-004",
        "name": "Kestrel Logistics",
        "industry": "logistics",
        "employees": 12500,
        "contract_size_usd": 720000,
        "public_reference": True,
        "approval_expires": (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat(),  # EXPIRED
        "deployed_products": ["platform", "enterprise_tier"],
    },
    {
        "customer_id": "cust-005",
        "name": "Meridian Energy",
        "industry": "energy",
        "employees": 22000,
        "contract_size_usd": 1200000,
        "public_reference": True,
        "approval_expires": (datetime.now(timezone.utc) + timedelta(days=320)).date().isoformat(),
        "deployed_products": ["platform", "enterprise_tier"],
    },
]

PRODUCTS_AND_TOPICS: dict[str, object] = {
    "products": [
        {"id": "platform", "name": "Core Platform", "tiers": ["team", "business", "enterprise"]},
        {"id": "enterprise_tier", "name": "Enterprise Tier", "features": ["BYOK", "audit_export", "priority_sla"]},
    ],
    "topics": [
        {"id": "encryption_at_rest", "label": "Encryption at rest"},
        {"id": "encryption_in_transit", "label": "Encryption in transit"},
        {"id": "key_management", "label": "Key management / BYOK"},
        {"id": "sso", "label": "Single sign-on"},
        {"id": "mfa", "label": "Multi-factor authentication"},
        {"id": "scim", "label": "SCIM provisioning"},
        {"id": "dr_bcp", "label": "Disaster recovery / BCP"},
        {"id": "incident_response", "label": "Incident response"},
        {"id": "ssdlc", "label": "Secure SDLC"},
        {"id": "pentest", "label": "Penetration testing"},
        {"id": "sbom", "label": "Software bill of materials"},
        {"id": "dpa", "label": "Data processing agreement"},
        {"id": "data_residency", "label": "Data residency"},
        {"id": "soc2", "label": "SOC 2"},
        {"id": "iso27001", "label": "ISO 27001"},
        {"id": "fedramp", "label": "FedRAMP"},
        {"id": "gdpr", "label": "GDPR"},
    ],
    "smes": [
        {"id": "sme-001", "name": "Priya Menon", "domains": ["encryption_at_rest", "encryption_in_transit", "key_management"]},
        {"id": "sme-002", "name": "Jordan Avery", "domains": ["dpa", "data_residency", "gdpr"]},
        {"id": "sme-003", "name": "Sam Okafor", "domains": ["incident_response", "pentest", "ssdlc"]},
        {"id": "sme-004", "name": "Alex Rivera", "domains": ["soc2", "iso27001", "fedramp"]},
    ],
}


def _write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, default=str))


def _write_markdown(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body)


# --------------------------------------------------------------------------
# Reference corpus JSON sidecars — uploaded to S3 by scripts/seed_s3_corpus.py.
# Primary sources (compliance/ + product-docs/) carry updated_at.
# Prior-RFP sidecars (prior-rfps/) carry approved_at.
#
# Staleness demo: three priors were approved BEFORE a corroborating primary
# source was last updated — they will be suppressed by Rule 1 (freshness
# suppression) and flagged by the staleness daemon (Rule 4).
# One prior was approved AFTER all corroborating sources — it stays live.
# --------------------------------------------------------------------------

CORPUS_PRIMARY: list[dict] = [
    # compliance/ — authoritative certificates and audit reports
    {
        "document_id": "soc2_cert_2025",
        "title": "SOC 2 Type II Certificate — 2025 Audit",
        "excerpt": (
            "SOC 2 Type II audit report covering trust principles Security, Availability, and "
            "Confidentiality. Audit period: 2024-10-01 to 2025-09-30. Auditor: independent "
            "third-party CPA firm. No qualified opinions. Next report due Q4 2026."
        ),
        "updated_at": "2025-10-01",
        "topic_ids": ["soc2"],
        "prefix": "compliance",
    },
    {
        "document_id": "iso27001_cert_2026",
        "title": "ISO 27001 Certificate — 2026 Renewal",
        "excerpt": (
            "ISO 27001:2022 certification. Certificate number: ISMS-2026-00142. "
            "Certification body: accredited external registrar. Valid through 2029-01-14. "
            "Scope: design, development, and operation of the SaaS platform."
        ),
        "updated_at": "2026-01-15",
        "topic_ids": ["iso27001"],
        "prefix": "compliance",
    },
    {
        "document_id": "fedramp_status_2026",
        "title": "FedRAMP Moderate — ATO Assessment Status",
        "excerpt": (
            "FedRAMP Moderate ATO assessment in progress as of 2026-03-01. Third-Party Assessment "
            "Organization (3PAO) engaged. No committed authorization date. Customers requiring "
            "FedRAMP Moderate should contact their Account Executive for timeline updates."
        ),
        "updated_at": "2026-03-01",
        "topic_ids": ["fedramp"],
        "prefix": "compliance",
    },
    # product-docs/ — authoritative technical documentation
    {
        "document_id": "encryption_whitepaper",
        "title": "Encryption Architecture Whitepaper",
        "excerpt": (
            "AES-256-GCM encryption at rest. AWS KMS managed keys; BYOK via customer CMK in "
            "customer's AWS account with key-policy binding. Key rotation: annual by default, "
            "on-demand via KMS API. TLS 1.2+ in transit with PFS (ECDHE cipher suites), HSTS "
            "one-year max-age, TLS 1.0/1.1 disabled."
        ),
        "updated_at": "2025-09-05",
        "topic_ids": ["encryption_at_rest", "encryption_in_transit", "key_management"],
        "prefix": "product-docs",
    },
    {
        "document_id": "sso_mfa_guide",
        "title": "SSO and MFA Integration Guide",
        "excerpt": (
            "SAML 2.0 and OIDC for customer-initiated SSO. OAuth 2.0 Authorization Code + PKCE "
            "for native integrations. SCIM 2.0 for automated provisioning. MFA factors: TOTP, "
            "WebAuthn/FIDO2, SMS (not recommended for high-assurance). Hardware keys required "
            "for admin roles in Enterprise tier."
        ),
        "updated_at": "2025-10-20",
        "topic_ids": ["sso", "mfa", "scim"],
        "prefix": "product-docs",
    },
    {
        "document_id": "dr_bcp_overview",
        "title": "Disaster Recovery and BCP Overview",
        "excerpt": (
            "Active-active deployment across two AWS regions. Synchronous replication of critical "
            "metadata; asynchronous replication < 60 s for bulk data. Target RTO: 4 hours. "
            "Target RPO: 15 minutes. DR exercises twice annually. Customer SLA: 99.9% monthly "
            "uptime (Enterprise tier). Backups: daily snapshots retained 90 days."
        ),
        "updated_at": "2025-11-10",
        "topic_ids": ["dr_bcp"],
        "prefix": "product-docs",
    },
]

# Three intentionally stale priors (approved_at < corroborating primary's updated_at).
# One fresh prior (approved_at > all corroborating sources).
CORPUS_PRIOR_RFPS: list[dict] = [
    {
        "document_id": "acme_financial_soc2_answer",
        "title": "Prior answer — Acme Financial: SOC 2 scope (STALE — cert updated after approval)",
        "excerpt": (
            "SOC 2 Type II audit covers Security, Availability, and Confidentiality trust "
            "principles. Most recent audit period ended June 2025. Report available under NDA."
        ),
        "approved_at": "2025-07-01",  # STALE: soc2_cert_2025 updated 2025-10-01
        "corroborated_by": ["compliance/soc2_cert_2025.json"],
        "topic_ids": ["soc2"],
    },
    {
        "document_id": "bluebird_encryption_answer",
        "title": "Prior answer — Bluebird Insurance: encryption at rest (STALE — whitepaper updated)",
        "excerpt": (
            "Customer data encrypted at rest using AES-256-GCM. Key management via AWS KMS. "
            "Enterprise BYOK available via CMK. Key rotation enabled."
        ),
        "approved_at": "2025-08-15",  # STALE: encryption_whitepaper updated 2025-09-05
        "corroborated_by": ["product-docs/encryption_whitepaper.json"],
        "topic_ids": ["encryption_at_rest", "key_management"],
    },
    {
        "document_id": "globex_fedramp_answer",
        "title": "Prior answer — Globex: FedRAMP status (STALE — status doc updated after approval)",
        "excerpt": (
            "FedRAMP Moderate ATO assessment is in progress. Third-party assessment organization "
            "has been engaged. No committed timeline for authorization."
        ),
        "approved_at": "2026-02-01",  # STALE: fedramp_status_2026 updated 2026-03-01
        "corroborated_by": ["compliance/fedramp_status_2026.json"],
        "topic_ids": ["fedramp"],
    },
    {
        "document_id": "northwind_sso_answer",
        "title": "Prior answer — Northwind Capital: SSO protocols (FRESH — approved after source)",
        "excerpt": (
            "SAML 2.0 and OIDC supported for customer-initiated SSO. SCIM 2.0 for automated "
            "user provisioning. OAuth 2.0 Authorization Code + PKCE for native integrations."
        ),
        "approved_at": "2026-01-15",  # FRESH: sso_mfa_guide updated 2025-10-20 (prior is newer)
        "corroborated_by": ["product-docs/sso_mfa_guide.json"],
        "topic_ids": ["sso", "scim"],
    },
]


def main() -> None:
    print(f"Generating synthetic data under {DATA}")

    # Incoming RFP
    _make_rfp_workbook(
        DATA / "incoming" / "sample_rfp_acmesec.xlsx",
        title="Acme Corp — Vendor Security Questionnaire",
        prospect="Acme Corp (fictional)",
        questions=INCOMING_RFP_QUESTIONS,
        with_answers=False,
    )
    print(f"  ✓ incoming/sample_rfp_acmesec.xlsx ({len(INCOMING_RFP_QUESTIONS)} questions)")

    # Historical RFPs
    for name, prospect, qa, win_loss in WON_RFPS:
        _make_rfp_workbook(
            DATA / "historical" / f"{name}.xlsx",
            title=f"{name.replace('_', ' ').title()}",
            prospect=prospect,
            questions=qa,
            with_answers=True,
            win_loss=win_loss,
        )
        print(f"  ✓ historical/{name}.xlsx ({len(qa)} Q&A, {win_loss})")

    # Seismic
    _write_json(DATA / "seismic" / "content_cards.json", SEISMIC_CARDS)
    print(f"  ✓ seismic/content_cards.json ({len(SEISMIC_CARDS)} cards)")

    # Gong
    _write_json(DATA / "gong" / "transcripts.json", GONG_TRANSCRIPTS)
    print(f"  ✓ gong/transcripts.json ({len(GONG_TRANSCRIPTS)} transcripts)")

    # Whitepapers
    for slug, body in WHITEPAPERS.items():
        _write_markdown(DATA / "whitepapers" / f"{slug}.md", body)
    print(f"  ✓ whitepapers/ ({len(WHITEPAPERS)} docs)")

    # Graph seed
    _write_json(DATA / "graph" / "customers.json", CUSTOMERS)
    _write_json(DATA / "graph" / "products_and_topics.json", PRODUCTS_AND_TOPICS)
    print(f"  ✓ graph/customers.json ({len(CUSTOMERS)} customers)")
    print("  ✓ graph/products_and_topics.json")

    # Reference corpus sidecars — primary sources
    for doc in CORPUS_PRIMARY:
        prefix = doc["prefix"]
        doc_id = doc["document_id"]
        payload = {k: v for k, v in doc.items() if k != "prefix"}
        _write_json(DATA / "corpus" / prefix / f"{doc_id}.json", payload)
    print(f"  ✓ corpus/compliance/ + corpus/product-docs/ ({len(CORPUS_PRIMARY)} primary sidecars)")

    # Reference corpus sidecars — prior RFPs (3 stale, 1 fresh for staleness demo)
    for doc in CORPUS_PRIOR_RFPS:
        _write_json(DATA / "corpus" / "prior-rfps" / f"{doc['document_id']}.json", doc)
    stale_count = sum(1 for d in CORPUS_PRIOR_RFPS if "STALE" in d["title"])
    fresh_count = len(CORPUS_PRIOR_RFPS) - stale_count
    print(f"  ✓ corpus/prior-rfps/ ({stale_count} stale, {fresh_count} fresh — staleness demo ready)")

    print("Done.")


if __name__ == "__main__":
    main()
